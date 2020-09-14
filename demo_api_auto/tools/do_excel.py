from openpyxl import load_workbook
from demo_api_auto.tools.read_config import ReadConfig
from demo_api_auto.tools.project_path import *


class DoExcel:
    @staticmethod
    def get_data(file_name):
        wb = load_workbook(file_name)
        mode = eval(ReadConfig.get_config(test_config_path, "MODE", "mode"))
        test_data = []
        for key in mode:
            sheet = wb[key]
            if mode[key] == "all":
                for i in range(2, sheet.max_row + 1):
                    row_data = {}
                    row_data["sheet_name"] = key
                    row_data["case_id"] = sheet.cell(i, 1).value
                    row_data["url"] = sheet.cell(i, 2).value
                    row_data["data"] = sheet.cell(i, 3).value
                    row_data["check_sql"] = sheet.cell(i, 4).value
                    row_data["title"] = sheet.cell(i, 5).value
                    row_data["http_method"] = sheet.cell(i, 6).value
                    row_data["expected"] = sheet.cell(i, 7).value
                    test_data.append(row_data)
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data["sheet_name"] = key
                    row_data["case_id"] = sheet.cell(case_id + 1, 1).value
                    row_data["url"] = sheet.cell(case_id + 1, 2).value
                    row_data["data"] = sheet.cell(case_id + 1, 3).value
                    row_data["check_sql"] = sheet.cell(case_id + 1, 4).value
                    row_data["title"] = sheet.cell(case_id + 1, 5).value
                    row_data["http_method"] = sheet.cell(case_id + 1, 6).value
                    row_data["expected"] = sheet.cell(case_id + 1, 7).value
                    test_data.append(row_data)

        return test_data

    @staticmethod
    def write_back(file_name, sheet_name, i, result, testresult):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 8).value = result
        sheet.cell(i, 9).value = testresult
        wb.save(file_name)
